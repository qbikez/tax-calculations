import os, sys, re
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from openpyxl import load_workbook
from helpers import convert_rate, convert_sheet, fiat_currencies, warsaw_timezone
from decimal import Decimal
from datetime import datetime

operations_to_skip = ["deposit", "withdrawal"]
operations_to_process = ["match", "fee"]

def calculate_tax():
    file_name = 'coinbase_pro.xlsx'

    if not os.path.exists(file_name):
        print(f'WARNING: Coinbase pro {file_name} doesnt exist. Skipping')
        return(None, None, None, None)

    workbook = load_workbook(filename=file_name)
    transactions = convert_sheet(workbook[workbook.sheetnames[0]])

    przychod_total = Decimal(0)
    koszt_total = Decimal(0)
    fiat_staking_total = Decimal(0)

    for row in transactions:
        if row['portfolio'] is None:
            continue

        type = row['type']
        if type in operations_to_skip:
            continue
        if type not in operations_to_process:
            raise Exception(f'Coinbase pro. Unknown transaction type {type}')

        asOfDate = row["time"]
        asOfDate = row["time"] if isinstance(row["time"], datetime) else datetime.strptime(row["time"], '%Y-%m-%d %H:%M:%S').astimezone(warsaw_timezone)
        trade_id = row["trade id"]
        amount = Decimal(str(row["amount"]))
        coin = row["amount/balance unit"]

        if type == 'fee' and coin in fiat_currencies:
            if amount >= 0:
                raise Exception(f"Positive fee for trade_id: {trade_id}")
            koszt_total -= round(convert_rate(asOfDate, amount, currency=coin), 2)

        if type != 'match' or coin not in fiat_currencies:
            continue
        pln = round(convert_rate(asOfDate, amount, currency=coin), 2)
        if pln > 0:
            przychod_total += pln
        else:
            koszt_total -= pln

    return ("Coinbase PRO", przychod_total, koszt_total, fiat_staking_total)