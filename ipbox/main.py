from functools import reduce
import os, traceback
from datetime import datetime, tzinfo
from typing import Iterable, List, NamedTuple
from azure.devops.v6_0.git.models import GitPullRequest
from azure.devops.v6_0.work_item_tracking.models import WorkItem
from dateutil.relativedelta import relativedelta
from openpyxl import cell, load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Fill
import pytz
from az import get_my_prs_from_repos, get_my_work_items_ids
from settings import excel_path, year, from_month, to_month, projects
import re

insensitive = re.compile('fix:?', re.IGNORECASE)

class ExcelRow(NamedTuple):
    merge_id: str
    work_item_id: str
    work_item_title: str
    pr_title: str
    closed_date: datetime
    url: str
    project: str

def build_excel_models(all_work_items: List[WorkItem], prs: List[GitPullRequest], project: str) -> Iterable[ExcelRow]:

    def obfuscate(title):
        return insensitive.sub('', title).strip().strip(':').strip()

    prs.sort(key=lambda x: x.closed_date)
    for pr in prs:
        work_items = list([w for w in all_work_items if w.id in pr.work_item_refs])
        if len(work_items) == 0:
            if len(pr.work_item_refs) > 0:
                print(f"Warn: {pr.id} has work items but could not find title")
            title = ""
            id = ""
        else:
            title = '; '.join([w.fields["System.Title"] for w in work_items])
            id = work_items[0].id
        print(f"task: {title} pr: {pr.title}")
        yield ExcelRow(pr.merge_id, id, obfuscate(title), obfuscate(pr.title), pr.closed_date, pr.url, pr.repository)

    for wi in all_work_items:
        if len([pr for pr in prs if wi.id in pr.work_item_refs]) == 0:
            print(f"task without PR: {wi.fields['System.Title']} {wi}")
            yield ExcelRow("", wi.id, wi.fields["System.Title"], "", wi.fields["Microsoft.VSTS.Common.ClosedDate"], wi.url, project)
            
def write_header(ws: Worksheet):
    header = ["PR Id", "Task Id", "Task title", "PR title", "Merged Date", "Days", "IsRND", "URL"]

    for i, h in enumerate(header):
        ws.cell(row=1, column=i + 1).value = h
        ws.cell(row=1, column=i + 1).font = Font(bold=True)

def is_excel_valid(month):
    if not os.path.exists(excel_path):
        return True
    wb = load_workbook(filename = excel_path)
    sheet_name = f'{year}-{month}'
    if sheet_name in wb.sheetnames and wb[sheet_name]["A2"].value is not None:
        return False
    return True

def write_excel(rows: Iterable[ExcelRow], sheet_name):
    exists = os.path.exists(excel_path)
    wb = load_workbook(filename = excel_path) if exists else Workbook()
    if not exists:
        wb.active.title = sheet_name
    ws: Worksheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    max_col = 5

    write_header(ws)
    for idx, row in enumerate(rows):
        row_id = idx + 2
        
        ws.cell(row=row_id, column=1).value = row.merge_id
        ws.cell(row=row_id, column=2).value = row.work_item_id
        ws.cell(row=row_id, column=3).value = row.work_item_title
        ws.cell(row=row_id, column=4).value = row.pr_title
        ws.cell(row=row_id, column=5).value = row.closed_date.astimezone(pytz.timezone("Poland")).replace(tzinfo=None)
        ws.cell(row=row_id, column=6).value = 0
        ws.cell(row=row_id, column=7).value = 0
        ws.cell(row=row_id, column=8).value = row.url
        ws.cell(row=row_id, column=9).value = row.project

    for col in range(1, max_col + 1):
         ws.column_dimensions[get_column_letter(col)].bestFit = True

    wb.save(excel_path)

total = []

for month in range(from_month, to_month + 1):
    if not is_excel_valid(month):
        print(f"Sheet for month {month} already has values. Skipping")
        continue

    start_date = datetime(year, month, 1, tzinfo=pytz.timezone("Poland"))
    end_date = start_date + relativedelta(months=1) - relativedelta(seconds=1)

    print(f"processing date range: {start_date} - {end_date}")
    
    try:
        excel_models = []
        for project in projects:
            print(f"> project: {project}")
            prs = list(get_my_prs_from_repos(start_date, end_date, project))
            work_items = list(get_my_work_items_ids(prs, start_date, end_date, project))

            print(f">  found {len(prs)} PRs and {len(work_items)} work items")

            excel_models += build_excel_models(work_items, prs, project)

        excel_models = ans = reduce(lambda acc, x: acc+[x] if x.merge_id not in [existing.merge_id for existing in acc] else acc, excel_models, [])
        total += excel_models
        write_excel(excel_models, f'{year}-{month}')
        print(f"Created {month}")
    except Exception:
        print(f"Failed on month {month}")
        traceback.print_exc()
        
write_excel(total, f'{year}-all')